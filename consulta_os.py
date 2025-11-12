#from finalizar_sem_reparo import aplicar_produto_entregue, aplicar_reparo_completo_remontagem, mudar_pra_ow, muda_tecnico_gspn, aplica_ag_custo_gspn
#from automacoes.cos.auto_cos import deletar_cos
from automacoes.cos.coletar_dados_cos import obter_os_correspondentes
#from automacoes.coletar_dados import fetch_os_data
from coletar_os_da_base.coletar_informações_os import coletar_informacoes_os
#from automacoes.montar_payloads import obter_dados_saw
import time
from login_gspn.cookies_manager import obter_cookies_validos_recentes
CAMINHO_ARQUIVO = r"C:\\Users\\Gestão MX\\Documents\\Copilot\\OS_consultar.txt"
"""Este módulo consulta a lista de ordens de serviço e imprime no console as informações de cada OS."""
import pandas as pd
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
import os
import re
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def formatar_saw(dados_saw):
    if not isinstance(dados_saw, dict):
        return {}
    saw_formatado = {}
    for i, (nome_saw, detalhes) in enumerate(dados_saw.items(), start=1):
        saw_formatado[f"SAW {i}"] = nome_saw
        saw_formatado[f"STATUS {i}"] = detalhes.get("status", "")
        saw_formatado[f"CATEGORIA {i}"] = detalhes.get("categoria", "")
    return saw_formatado

def gerar_nome_arquivo(diretorio="C:\\Users\\Gestão MX\\Documents\\Consultas OS", prefixo="consulta"):
    hoje_str = datetime.now().strftime("%d.%m.%Y")
    contador = 1
    while True:
        nome = f"{prefixo} {hoje_str}-{contador:03}.xlsx"
        caminho = os.path.join(diretorio, nome)
        if not os.path.exists(caminho):
            return caminho
        contador += 1

def gerar_excel_os(lista_os_list):
    lista_expandidas = []

    for os_container in lista_os_list:
        os_id, dados = next(iter(os_container.items()))
        base = dados.copy()

        # Trata o campo SAW
        saw_data = base.pop("dados_saw", "Não possui SAW")
        if isinstance(saw_data, dict):
            saw_formatado = formatar_saw(saw_data)
            base.update(saw_formatado)
        else:
            base["dados_saw"] = saw_data

        lista_expandidas.append(base)

    df = pd.DataFrame(lista_expandidas)
    #df.columns = df.columns.str.strip()  # remove espaços

    if "fim_garantia" in df.columns:
        df["fim_garantia"] = df["fim_garantia"].astype(str).str.replace(r"\(L\)", "", regex=True)
        df["fim_garantia"] = pd.to_datetime(df["fim_garantia"], format="%d/%m/%Y", errors='coerce')
    else:
        df["fim_garantia"] = pd.NaT

    # Geração de Excel com openpyxl para aplicar estilos
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordens de Serviço"

    # Escreve os dados
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Estilos
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # ruim
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # neutro

    hoje = datetime.now().date()
    cinco_dias = hoje + timedelta(days=5)

    # Cabeçalhos
    header = [cell.value for cell in ws[1]]

    col_status_garantia = header.index("status_garantia_cos") + 1
    col_fim_garantia = header.index("fim_garantia") + 1
    col_dados_saw = header.index("dados_saw") + 1 if "dados_saw" in header else None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        status = row[col_status_garantia - 1].value
        print(status)
        fim = row[col_fim_garantia - 1].value
        saw = row[col_dados_saw - 1].value if col_dados_saw else ""

        # Marcar como RUIM
        if status == "Garantia Samsung":
            try:
                if isinstance(fim, datetime):
                    fim_date = fim.date()
                else:
                    fim_date = datetime.strptime(str(fim), "%Y-%m-%d").date()

                if fim_date <= cinco_dias:
                    for cell in row:
                        cell.fill = red_fill
            except Exception as e:
                print(f"Erro ao processar data '{fim}': {e}")

        # Marcar como NEUTRO (se tiver dados de SAW válidos)
        elif any("SAW 1" in str(cell.value) for cell in row):
            for cell in row:
                cell.fill = gray_fill
    from openpyxl.utils import get_column_letter

    # Ajuste automático da largura das colunas
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2  # +2 pra dar um respiro
        ws.column_dimensions[col_letter].width = adjusted_width
    # Salvar arquivo com nome único
    nome_arquivo = gerar_nome_arquivo()
    wb.save(nome_arquivo)
    print(f"Arquivo salvo como: {nome_arquivo}")


def fechar_lista_de_os(cookies=None):

    print("Iniciando consulta de lista de OS...")
    fallhas = []
    with open(CAMINHO_ARQUIVO, "r", encoding="utf-8") as file:
        os_list = [line.strip() for line in file.readlines() if line.strip()]
    total=[]
    for os in os_list:
        #deletar = deletar_cos(os)
        #os = obter_os_correspondentes(os)
        #dados= fetch_os_data(os)
        #fechamento = mudar_pra_ow(dados)
        #fechamento = aplica_ag_custo_gspn(os)
        #fechamento = muda_tecnico_gspn(os)
        try:
            fechamento = coletar_informacoes_os(os, cookies)
        except:
            fallhas.append(os)
            continue

        #fechamento = obter_dados_saw(os, cookies)
        #fechamento = aplicar_reparo_completo_remontagem(os)
        #if fechamento == []:
            #fechamento = "Não há SAW"
        if fechamento:
            fechamento = {os: fechamento}
            
            print(f"Consulta da OS {os} realizado com sucesso.")
            #print(fechamento)
            total.append(fechamento)
            #print(total)
            #time.sleep(2)
        else:
            print(f"Erro ao consultar a OS {os}.")
            fallhas.append(os)
            continue
        """print('Aplicando o produto entregue...')
        produto_entregue = aplicar_produto_entregue(os)
        if produto_entregue:
            print(f"Produto entregue aplicado na OS {os} com sucesso.")
            print("Deletando OS COS...")
            deletar = deletar_cos(os)
            if deletar:
                print(f"OS COS {os} deletada com sucesso.")
            else:
                print(f"Erro ao deletar a OS COS {os}.")
        else:
            print(f"Erro ao aplicar o produto entregue na OS {os}.")
            fallhas.append(os)"""
    if fallhas:
        print("As seguintes OSs falharam no fechamento:")
        for os in fallhas:
            print(os)

    #Gera uma planilha com os dados
    sucesso = gerar_excel_os(total)
    if sucesso:
        print("Planilha gerada com sucesso!")
    else:
        print("Erro ao gerar a planilha.")
    for fechamento in total:
        print('------Dados da OS-------')
        #print(fechamento)
        
        # Se fechamento ainda for um dicionário com a chave sendo o número da OS:
        if isinstance(fechamento, dict) and len(fechamento) == 1:
            os_numero, dados = next(iter(fechamento.items()))
            #print(f'OS: {os_numero}')
        else:
            dados = fechamento  # Caso seja apenas o dicionário direto

        for chave, valor in dados.items():
            if chave == 'dados_saw' and isinstance(valor, dict):
                print('Dados de SAW:')
                for saw_id, saw_info in valor.items():
                    print(f'  {saw_id}:')
                    for campo, desc in saw_info.items():
                        print(f'    {campo}: {desc}')
            else:
                print(f'{chave}: {valor}')
        




if __name__ == "__main__":
    cookies = obter_cookies_validos_recentes()
    fechar_lista_de_os(cookies)
    #aplicar_reparo_completo_remontagem("4172771585")
    print("Consulta da lista de OS concluído.")